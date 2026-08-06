"""Excel 解析：每个 sheet 转 Markdown 表格，超大表按行分块。"""
from __future__ import annotations

from pathlib import Path

from app.core.exceptions import DocumentParseError
from app.ingestion.parsers.base import BaseParser, ParsedBlock, ParsedDocument

MAX_ROWS_PER_BLOCK = 40


class ExcelParser(BaseParser):
    suffixes = (".xlsx", ".xls", ".csv")

    def parse(self, path: Path) -> ParsedDocument:
        doc = ParsedDocument(meta={"source": path.name})
        if path.suffix.lower() == ".csv":
            self._parse_csv(path, doc)
            return doc

        try:
            from openpyxl import load_workbook
        except ImportError as e:  # pragma: no cover
            raise DocumentParseError("未安装 openpyxl") from e
        try:
            wb = load_workbook(str(path), read_only=True, data_only=True)
        except Exception as e:  # noqa: BLE001
            raise DocumentParseError(f"Excel 打开失败: {e}")

        for ws in wb.worksheets:
            rows = [[("" if c is None else str(c).strip()) for c in r]
                    for r in ws.iter_rows(values_only=True)]
            rows = [r for r in rows if any(r)]
            if not rows:
                continue
            doc.blocks.append(ParsedBlock(text=f"# 工作表：{ws.title}", block_type="heading",
                                          heading_level=1))
            header = rows[0]
            for start in range(1, len(rows), MAX_ROWS_PER_BLOCK):
                part = rows[start:start + MAX_ROWS_PER_BLOCK]
                lines = ["| " + " | ".join(header) + " |",
                         "|" + "---|" * len(header)]
                lines += ["| " + " | ".join(r + [""] * (len(header) - len(r))) + " |"
                          for r in part]
                doc.blocks.append(ParsedBlock(text="\n".join(lines), block_type="table",
                                              meta={"sheet": ws.title}))
        wb.close()
        if not doc.blocks:
            raise DocumentParseError("Excel 无有效数据")
        return doc

    @staticmethod
    def _parse_csv(path: Path, doc: ParsedDocument) -> None:
        import csv
        with open(path, "r", encoding="utf-8-sig", errors="ignore") as f:
            rows = [r for r in csv.reader(f) if any(x.strip() for x in r)]
        if not rows:
            raise DocumentParseError("CSV 无有效数据")
        header = rows[0]
        for start in range(1, len(rows), MAX_ROWS_PER_BLOCK):
            part = rows[start:start + MAX_ROWS_PER_BLOCK]
            lines = ["| " + " | ".join(header) + " |", "|" + "---|" * len(header)]
            lines += ["| " + " | ".join(r) + " |" for r in part]
            doc.blocks.append(ParsedBlock(text="\n".join(lines), block_type="table"))
